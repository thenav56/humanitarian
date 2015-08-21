"""Module for extracting data from dropbox and aggregating"""
#TODO: better way for specifyign which folder to pull from (latest?)
##how to handle names?
##qualify sheets as being valid before cycling or skip invalid ones?
##pull from rules based text file
##log put on dbox, make better
##test individual cleaning
##upload new version

import dropbox
import clean
import os
import cStringIO
import re
from openpyxl import load_workbook
import openpyxl.writer.excel as wrtex


#dropbox setup
db_key = os.environ['db_key']
db_secret = os.environ['db_secret']
db_access = os.environ['db_access']
client = dropbox.client.DropboxClient(db_access)

DB_PATH = '/2015 Nepal EQ/04 IM/Reporting/Incoming_submissions_Z/'
TEST_FILE = '/test_sheet.xlsx'

def iterate_reports():
    """cycle through all reports contained in dbox directory"""

    meta = client.metadata(DB_PATH, list=True)
    file_list = [str(f['path']) for f in meta['contents'] if re.search('xls|xlsx$',str(f))]
    exclude = [
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/3w-Habitat for Humanity - 8-14-2015 dwld.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/2015 08 20_ CARE_Shelter 4Ws_final.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/ADRA - 20082015.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/Caritas Switzerland - 17082015.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/CDRA Nepal.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/Child Space 19 August 2015.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/Copy of Cesvi 18082015_sheltercluster.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/Copy of reportingtemplate_sheltercluster_v4.0_5.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/CRS-Caritas - 20082015.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/Germany-GIZ.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/IOM 18 August 2015.xlsx',
    '/2015 Nepal EQ/04 IM/reporting/Incoming_submissions_Z/ISR-NEPAL 4W Templete.xlsx'
    ]

    for v in exclude:
        file_list.remove(v)

    #file_list = [DB_PATH+"/Batas Foundation.xlsx/"]
    for f in file_list:
        #pull down workbook from specified directory
        print "pulling! " + f
        wb_current = pull_wb(f)

        #check to see if properly formatted
        if wb_format(wb_current):
            #clean the workboook
            print "cleaning " + f
            clean_file(wb_current, f)

        else:
            #put in malformatted folder
            print 'malformatted ' + f
            path = DB_PATH + '/old_format_or_incorrect/' + f.rsplit('/', 1)[1]
            #send_wb(path, wb_current)

def send_wb(path, wb):
    print 'sending! ' + path
    client.put_file(path, wrtex.save_virtual_workbook(wb))


def wb_format(wb):
    """check to see if a report is correct and in the new report format"""
    must_contain = ['Guidance Note', 'Distributions', 'Training', 'Reference']

    match_count = 0
    for s in wb.worksheets:
        if s.title in must_contain:
            match_count+=1

    if match_count < 4:
        return False
    else:
        return True

def clean_file(wb, path): 
    """cycle through a report and apply cleaning algorithms"""
    
    #get our two sheets
    db = wb.get_sheet_by_name('Distributions')
    ref = wb.get_sheet_by_name('Reference')

    #setup log
    rname =  path.rsplit('/', 1)[1]
    report_line = '***Report for ' + rname
    report_a_log(report_line, rname)


    #####do edit stuff
    #algos return db, ref, message

    #algo1
    db, ref, message = clean.algo1(db,ref) 
    report_a_log(message, rname)

    #algo2
    db, ref, message = clean.algo2(db,ref)
    report_a_log(message, rname)

    #algo3
#    db, ref, message = clean.algo3(db,ref)
#    report_a_log(message, rname)

    #algo4
    db, ref, message = clean.algo4(db,ref)
    report_a_log(message, rname)

    #algo5
    db, ref, message = clean.algo5(db,ref)
    report_a_log(message, rname)

    #algo6
    db, ref, message = clean.algo6(db,ref)
    report_a_log(message, rname)

    #algo7
    db, ref, message = clean.algo7(db,ref)
    report_a_log(message, rname)

    #algo8
    db, ref, message = clean.algo8(db,ref)
    report_a_log(message, rname)

    #algo9
    db, ref, message = clean.algo9(db,ref)
    report_a_log(message, rname)

    #algo10
    db, ref, message = clean.algo10(db,ref)
    report_a_log(message, rname)

    #algo11
    db, ref, message = clean.algo11(db,ref)
    report_a_log(message, rname)

    #algo12
    db, ref, message = clean.algo12(db,ref)
    report_a_log(message, rname)

    #algo13
    db, ref, message = clean.algo13(db,ref)
    report_a_log(message, rname)

    #algo14
    db, ref, message = clean.algo14(db,ref)
    report_a_log(message, rname)

    #algo15
    db, ref, message = clean.algo15(db,ref)
    report_a_log(message, rname)

    #algo16
    db, ref, message = clean.algo16(db,ref)
    report_a_log(message, rname)

    #algo17
    db, ref, message = clean.algo17(db,ref)
    report_a_log(message, rname)

    #algo18
    db, ref, message = clean.algo18(db,ref)
    report_a_log(message, rname)

    #algo19
    db, ref, message = clean.algo19(db,ref)
    report_a_log(message, rname)


    #dummy empty log to send to finalize logging
    report_a_log(' ','text')

    #upload with name of file at end
    #we need to upload the new version!!!!!!!!
    send_wb(DB_PATH + '/edited/' + path.rsplit('/', 1)[1], wb)
    print 'uploaded! ' + path

report_recvd = False
current_path = ''
current_log = []
old_path = ""

def report_a_log(log_value, path):
    """write out contents for a given log - ends if a new path is given"""
    #todo: this is gross
    global report_recvd
    global current_path
    global current_log
    global old_path

    #if module is starting and we haven't logged anything
    if not report_recvd:
        current_path = path
        report_recvd = True
        current_log.append(log_value)
        current_log.append('')
        old_path = path
    
    #if we are recieving a new path
    elif current_path != path:
        current_path = path
        
        #write out
        with open('/Users/ewanog/code/nepal-earthquake/shelter/etl/etl/logs/' 
            + old_path + '.txt', 'w') as f:
            for log in current_log:
                f.write(str(log)+'\n')
        f.close()

        #create new current_log
        current_log = [log_value]
        current_log.append('')
        old_path = path

    else:
        current_log.append(log_value)
        current_log.append('')

def find_in_header(sheet, find_val):
    """find the coordinate of a value in header (assumes header is in row 1)"""
    for row in sheet.iter_rows('A1:' + find_last_value(sheet,'A','r')):
        for cell in row:
            if cell.value == find_val:
                return cell.column

    #if we haven't returned anything yet
    return None

def colvals_notincol(sheet_val,col_val,sheet_ref,col_ref):
    """return values from a column that are NOT in a reference column"""
    not_in = []
    to_search = []

    #create an array from sheet_ref with values to be searched (as opposed to nested loops)
    #iter_rows syntax: sheet.iter_rows('A1:A2')
    for row in sheet_ref.iter_rows(col_ref + "2:" + 
        find_last_value(sheet_ref, col_ref, 'c')):

        for cell in row:               
                try:
                    to_search.append(str(cell.value.encode('utf8')))
                except:
                    to_search.append(str(cell.value))

    #now search through vals and see if they're present
    for row in sheet_val.iter_rows(col_val + "2:" + 
        find_last_value(sheet_val, col_val, 'c')):

        for cell in row:
            if str(cell.value) not in to_search:
                try:
                    not_in.append(str(cell.value.encode('utf8')))
                except:
                    not_in.append(str(cell.value))

    return not_in
             


def find_last_value(sheet, start_location, r_or_c):
    """find position of last value in a given row or column"""
    #extract form r_or_c if we should be iterating a row or column
    
    row_it = 0
    col_it = 0
    if r_or_c == 'c':
        row_it = 1
    elif r_or_c == 'r':
        col_it = 1
    else:
        raise Exception("r_or_c must be r or c!")
    
    #look for a cell without value, and if we find a blank, traverse 100 more
    #to make sure there's no blanks
    blank_count = 0
    found = False
    cur_cell = sheet[start_location+'1']
    while not found:
        cur_cell = cur_cell.offset(row = row_it, column = col_it)

        if not cur_cell.value:
            if blank_count == 0:
                #coord of a cell step back 1
                last_found = cur_cell.offset(row = -row_it, column= -col_it).coordinate
                blank_count+=1

            elif blank_count == 100:
                found = True
            else:
                blank_count+=1
        else:
            blank_count=0

    return last_found


def pull_wb(location):
    """return an excel pulled from dropbox"""

    in_mem_file = pull_file(location)

    wb = load_workbook(in_mem_file)
    print "pulled! " + str(wb.get_sheet_names())  
    in_mem_file.close()

    return wb

def pull_file(location):
    """pull a file from dropbox"""
    to_ret = cStringIO.StringIO()

    with client.get_file(location) as f:
        to_ret.write(f.read())
    f.close()

    return to_ret

def test():
    return load_workbook('/Users/ewanog/code/nepal-earthquake/shelter/etl/clean_test.xlsx', data_only=True)

if __name__ == '__main__':
    iterate_reports()